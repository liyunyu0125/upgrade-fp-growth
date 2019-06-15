 ##-*- coding: utf-8 -*-
import openpyxl
import xlwt

def unique_index(L,f):
        ##L表示列表， i表示索引值，v表示values，f表示要查找的元素
        return [i for (i,v) in enumerate(L) if v==f]
    
wb = openpyxl.load_workbook(r'C:\Users\liyun\Desktop\毕设\数据\programming\data.xlsx')
sheet1 = wb.get_sheet_by_name("Sheet1")
originalrow=sheet1.max_row
indicatorlist=[]
qcnum=[]
agvnum=[]
armgnum=[]
task=[]
policy=[]
key=[]
agvnot=[]
agvyes=[]
sumarmg=[]
qcaverage=[]
agvaverage=[]
armgarveage=[]
qceffi=[]
agveffi=[]
armgeffi=[]
timecost=[]

for rows in range(2,originalrow):
    indicator0=[]
    indicator1=[]
    indicator=[]
    indicator0.append(sheet1.cell(rows,17).value[unique_index(sheet1.cell(rows,17).value,'.')[0]-1:unique_index(sheet1.cell(rows,17).value,',')[8]])#第1个‘.’出现的位置前一位到第9个‘，’出现的位置前
    indicator1.append(sheet1.cell(rows,17).value[unique_index(sheet1.cell(rows,17).value,'[')[2]+1:unique_index(sheet1.cell(rows,17).value,',')[unique_index(unique_index(sheet1.cell(rows,17).value,','),unique_index(sheet1.cell(rows,17).value,'[')[2]-2)[0]+9]])#第3个‘[’出现的位置后一位到第9个‘，’出现的位置前
    indicator.append([indicator0,indicator1])
    indicatorlist.append(indicator)
    qcnum.append(float(sheet1.cell(rows,1).value))
    agvnum.append(float(sheet1.cell(rows,2).value))
    armgnum.append(float(sheet1.cell(rows,3).value))
    task.append(float(sheet1.cell(rows,4).value))
    policy.append(float(sheet1.cell(rows,5).value))
    key.append(float(sheet1.cell(rows,6).value))
    agvnot.append(float(sheet1.cell(rows,7).value))
    agvyes.append(float(sheet1.cell(rows,8).value))
    sumarmg.append(float(sheet1.cell(rows,9).value))
    qcaverage.append(float(sheet1.cell(rows,10).value))
    agvaverage.append(float(sheet1.cell(rows,11).value))
    armgarveage.append(float(sheet1.cell(rows,12).value))
    qceffi.append(float(sheet1.cell(rows,13).value))
    agveffi.append(float(sheet1.cell(rows,14).value))
    armgeffi.append(float(sheet1.cell(rows,15).value))    
    timecost.append(float(sheet1.cell(rows,16).value))
#indicatorlist[第几行，][0][key][0]

book = xlwt.Workbook()
sheet = book.add_sheet('data analysis', cell_overwrite_ok=True)

colume_name=['qc数量','agv数量','armg数量','任务数量','指派策略选择','key','AGV不归位距离','AGV归位距离','sum(ARMG)','QC平均利用率','AGV平均利用率','ARMG平均利用率','QC平均效率','AGV平均效率','ARMG平均效率','timecost','平均AGV到达待选任务接箱点的空驶距离','平均AGV指派待选任务所需的重驶距离','3.	平均AGV到达待选任务接箱点时刻与其他AGV中最快到达该点时刻差','平均待选任务的装-1/卸1类型','平均待选任务所在ARMG的相对剩余工作量','平均待选任务的起重机交箱耗时','平均待选任务所在QC的平均延迟','平均待选任务指派给当前AGV的dualcycle可能性','平均选任务的duetime']
row = 0
for item in range(len(colume_name)):
    sheet.write(row, item, colume_name[item])
    
for rowNum in range(0,118):   
    for key in range(2):
        #count为列，一行9位write进excel
        element=[]
        count=0
        finddot=1
        while finddot !=0:
            if indicatorlist[rowNum][0][key][0].find(',',finddot,indicatorlist[rowNum][0][key][0].rfind(',')+1) != -1:
                element.append(indicatorlist[rowNum][0][key][0].find(',',finddot,indicatorlist[rowNum][0][key][0].rfind(',')+1))
                finddot=1 + indicatorlist[rowNum][0][key][0].find(',',finddot,indicatorlist[rowNum][0][key][0].rfind(',')+1)
            else:
                finddot=0
            
        
        while count <= len(element):
            if count==0:
                sheet.write(2*rowNum+1+key,count+16,float(indicatorlist[rowNum][0][key][0][:element[count]]))
            elif count==len(element):
                sheet.write(2*rowNum+1+key,count+16,float(indicatorlist[rowNum][0][key][0][element[count-1]+1:]))
            else:
                sheet.write(2*rowNum+1+key,count+16,float(indicatorlist[rowNum][0][key][0][element[count-1]+1:element[count]]))
            count=count+1
            
        sheet.write(2*rowNum+1+key,0,qcnum[rowNum+1])
        sheet.write(2*rowNum+1+key,1,agvnum[rowNum+1])
        sheet.write(2*rowNum+1+key,2,armgnum[rowNum+1])
        sheet.write(2*rowNum+1+key,3,task[rowNum+1])
        sheet.write(2*rowNum+1+key,4,policy[rowNum+1])
        sheet.write(2*rowNum+1+key,5,key)
        sheet.write(2*rowNum+1+key,6,agvnot[rowNum+1])
        sheet.write(2*rowNum+1+key,7,agvyes[rowNum+1])
        sheet.write(2*rowNum+1+key,8,sumarmg[rowNum+1])
        sheet.write(2*rowNum+1+key,9,qcaverage[rowNum+1])
        sheet.write(2*rowNum+1+key,10,agvaverage[rowNum+1])
        sheet.write(2*rowNum+1+key,11,armgarveage[rowNum+1])
        sheet.write(2*rowNum+1+key,12,qceffi[rowNum+1])
        sheet.write(2*rowNum+1+key,13,agveffi[rowNum+1])
        sheet.write(2*rowNum+1+key,14,armgeffi[rowNum+1]) 
        sheet.write(2*rowNum+1+key,15,timecost[rowNum+1])
    
book.save(r'C:\Users\liyun\Desktop\毕设\数据\programming\New Data.xls')
##wb.save(r'C:\Users\Liyun Yu\Desktop\毕设\数据\test\New Data.xlsx')  # 在字符串前加r，声明为raw字符串，这样就不会处理其中的转义了。否则，可能会报错
